"""
AuditRAM - Text search + red bounding box overlay tool
Supports: PDF, Image (png/jpg), Excel (.xlsx), Word (.docx - via optional docx2pdf)
Outputs: For PDF -> new PDF with red rectangle overlays (original unchanged)
         For Image -> new image with red rectangle overlays
         For Excel -> copy workbook with red cell borders on matching cells (original unchanged)
         For DOCX -> attempts conversion to PDF (needs docx2pdf); otherwise produces a text report of matches.

Usage (command-line):
    python main.py --input /path/to/file.pdf --text "Invoice Number" --output /path/to/out.pdf

Optional: run with Streamlit UI:
    pip install streamlit
    streamlit run main.py -- --ui

Dependencies (see requirements.txt). Some features require system installs:
 - Tesseract-OCR for image OCR (pytesseract). Install separately (e.g., on Ubuntu: sudo apt install tesseract-ocr)
 - docx2pdf is optional and works on Windows/MS Word or macOS with Word.

Author: Generated for Madhurima Das (AuditRAM assignment)
"""
import argparse
import os
import sys
import shutil
from pathlib import Path

def is_pdf(path): return str(path).lower().endswith('.pdf')
def is_image(path): return str(path).lower().endswith(('.png','.jpg','.jpeg','.tiff','.bmp'))
def is_xlsx(path): return str(path).lower().endswith('.xlsx')
def is_docx(path): return str(path).lower().endswith('.docx')

def ensure_output_dir(path):
    out = Path(path).parent
    out.mkdir(parents=True, exist_ok=True)

def process_pdf(input_path, search_text, output_path):
    try:
        import fitz  # PyMuPDF
    except Exception as e:
        print("PyMuPDF (fitz) is required for PDF processing. Install with: pip install pymupdf")
        raise

    doc = fitz.open(input_path)
    # We'll create a copy and draw rectangles on it
    for page_num in range(len(doc)):
        page = doc[page_num]
        text_instances = page.search_for(search_text, quads=False)
        # search_for returns list of rects
        for inst in text_instances:
            # draw rectangle - red, no fill
            shape = page.new_shape()
            shape.draw_rect(inst)
            shape.finish(color=(1, 0, 0), fill=None, width=1)  # RGB 1,0,0
            shape.commit()
    doc.save(output_path)
    doc.close()
    print(f"Saved PDF with overlays to: {output_path}")

def process_image(input_path, search_text, output_path):
    try:
        from PIL import Image, ImageDraw
        import pytesseract
    except Exception as e:
        print("Pillow and pytesseract are required for image processing. Install with: pip install pillow pytesseract")
        raise

    img = Image.open(input_path).convert("RGB")
    # Use pytesseract to get box data
    data = pytesseract.image_to_data(img, output_type=pytesseract.Output.DICT)
    draw = ImageDraw.Draw(img)
    n_boxes = len(data['level'])
    matches = 0
    st = search_text.lower()
    for i in range(n_boxes):
        word = str(data['text'][i]).strip()
        if word=="":
            continue
        if st in word.lower() or word.lower() in st:
            (x, y, w, h) = (data['left'][i], data['top'][i], data['width'][i], data['height'][i])
            # draw red rectangle (unfilled)
            draw.rectangle([x, y, x+w, y+h], outline=(255,0,0), width=2)
            matches += 1
    img.save(output_path)
    print(f"Saved image with overlays to: {output_path}. Matches highlighted: {matches}")

def process_xlsx(input_path, search_text, output_path):
    try:
        import openpyxl
        from openpyxl.styles import Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as e:
        print("openpyxl is required for Excel processing. Install with: pip install openpyxl")
        raise

    wb = openpyxl.load_workbook(input_path)
    thin_red = Side(border_style="thin", color="FF0000")
    border = Border(left=thin_red, right=thin_red, top=thin_red, bottom=thin_red)
    matches = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                if str(search_text).lower() in str(val).lower():
                    # apply red border to the cell (on a copy only)
                    cell.border = border
                    matches += 1
    wb.save(output_path)
    print(f"Saved Excel copy with red borders to: {output_path}. Matches: {matches}")

def process_docx(input_path, search_text, output_path):
    # Try converting to PDF using docx2pdf if available
    try:
        from docx2pdf import convert
        # Convert to a temporary PDF and then use PDF processing
        tmp_pdf = str(Path(output_path).with_suffix('.tmp_convert.pdf'))
        convert(input_path, tmp_pdf)
        from shutil import copyfile
        # Now process the generated PDF
        process_pdf(tmp_pdf, search_text, output_path)
        os.remove(tmp_pdf)
    except Exception as e:
        # Fallback: extract text occurrences and write a report (no coordinates)
        try:
            import docx
            doc = docx.Document(input_path)
            matches = []
            for i, para in enumerate(doc.paragraphs):
                if search_text.lower() in para.text.lower():
                    matches.append((i+1, para.text.strip()))
            report_path = str(Path(output_path).with_suffix('.txt'))
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write(f"Search report for '{search_text}' in {input_path}\\n\\n")
                for pnum, text in matches:
                    f.write(f"Paragraph {pnum}: {text}\\n")
            print(f"Could not convert DOCX to PDF automatically. Wrote a textual report to: {report_path}")
        except Exception as e2:
            print("docx2pdf or python-docx not available. Install docx2pdf (needs Word) for full overlay support, or python-docx for textual reports.")
            raise

def main():
    parser = argparse.ArgumentParser(description="AuditRAM text locator + overlay tool")
    parser.add_argument('--input', '-i', required=False, help="Input file path")
    parser.add_argument('--text', '-t', required=False, help="Text to search (case-insensitive)")
    parser.add_argument('--output', '-o', required=False, help="Output file path")
    parser.add_argument('--ui', action='store_true', help="Run Streamlit UI")
    args, unknown = parser.parse_known_args()

    # Support streamlit entrypoint: streamlit passes args after --, so check unknown
    if '--ui' in sys.argv or args.ui:
        try:
            import streamlit as st
        except:
            print("Streamlit not installed. Install with: pip install streamlit")
            return
        run_streamlit_app()
        return

    if not args.input or not args.text or not args.output:
        print("Missing arguments. Example:")
        print("  python main.py --input /path/file.pdf --text \"Invoice\" --output /path/out.pdf")
        return

    input_path = args.input
    search_text = args.text
    output_path = args.output
    ensure_output_dir(output_path)

    if is_pdf(input_path):
        process_pdf(input_path, search_text, output_path)
    elif is_image(input_path):
        process_image(input_path, search_text, output_path)
    elif is_xlsx(input_path):
        # create output copy name if not specified same extension
        process_xlsx(input_path, search_text, output_path)
    elif is_docx(input_path):
        process_docx(input_path, search_text, output_path)
    else:
        print("Unsupported file type. Supported: PDF, PNG/JPG, XLSX, DOCX")

# Simple Streamlit UI
def run_streamlit_app():
    import streamlit as st
    st.title("AuditRAM - Text search & bounding box overlay")
    uploaded_file = st.file_uploader("Upload a file (PDF, image, DOCX, XLSX)", type=['pdf','png','jpg','jpeg','docx','xlsx'])
    search_text = st.text_input("Text to search", "")
    if uploaded_file is not None and search_text.strip() != "":
        with st.spinner("Processing..."):
            tmp_dir = Path("temp_uploads")
            tmp_dir.mkdir(exist_ok=True)
            in_path = tmp_dir / uploaded_file.name
            with open(in_path,'wb') as f:
                f.write(uploaded_file.getbuffer())
            out_name = in_path.stem + "_overlay"
            if in_path.suffix.lower() == ".pdf":
                out_path = tmp_dir / (out_name + ".pdf")
            elif in_path.suffix.lower() in [".png",".jpg",".jpeg"]:
                out_path = tmp_dir / (out_name + in_path.suffix.lower())
            elif in_path.suffix.lower() == ".xlsx":
                out_path = tmp_dir / (out_name + ".xlsx")
            elif in_path.suffix.lower() == ".docx":
                out_path = tmp_dir / (out_name + ".pdf")
            else:
                st.error("Unsupported file type")
                return
            try:
                # Call the same processing functions
                if in_path.suffix.lower() == ".pdf":
                    process_pdf(str(in_path), search_text, str(out_path))
                    st.download_button("Download Result PDF", data=open(out_path,'rb'), file_name=out_path.name)
                elif in_path.suffix.lower() in [".png",".jpg",".jpeg"]:
                    process_image(str(in_path), search_text, str(out_path))
                    st.image(str(out_path))
                    st.download_button("Download Result Image", data=open(out_path,'rb'), file_name=out_path.name)
                elif in_path.suffix.lower() == ".xlsx":
                    process_xlsx(str(in_path), search_text, str(out_path))
                    st.write("Excel copy with red borders saved.")
                    st.download_button("Download Excel", data=open(out_path,'rb'), file_name=out_path.name)
                elif in_path.suffix.lower() == ".docx":
                    process_docx(str(in_path), search_text, str(out_path))
                    st.write("DOCX processed (check output).")
            except Exception as e:
                st.error(f"Processing failed: {e}")

if __name__ == '__main__':
    main()
